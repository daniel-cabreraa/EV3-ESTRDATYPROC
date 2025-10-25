import datetime as dt
from datetime import timedelta
import openpyxl as pyxl
from openpyxl.styles import Font, Border, Side, Alignment
import sqlite3
from sqlite3 import Error
import os
import sys

nombresTurnos = {"M": "Matutino", "V": "Vespertino", "N": "Nocturno"}
hoy = dt.datetime.now()

def agregarCliente():
    print("Escribe '0' en cualquier campo para cancelar la operación.")
    while True:
        nombre = input("Ingresa el nombre(s): ")
        if nombre == "" or nombre.strip() == "":
            print("ⓘ El nombre no puede estar vacío.")
            continue
        elif nombre == "0":
            return
        break
    while True:
        apellidos = input("Ingresa los apellidos: ")
        if apellidos == "" or apellidos.strip() == "":
            print("ⓘ Los apellidos no pueden estar vacíos.")
            continue
        elif apellidos == "0":
            return
        break
    try:
        with sqlite3.connect("estado.db") as conn:
            miCursor = conn.cursor()
            datosCliente = (nombre, apellidos)
            miCursor.execute("INSERT INTO clientes (nombre, apellidos) VALUES(?,?)", datosCliente)
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    else:
        print("✓ Cliente agregado con éxito.\n")

def mostrarClientes():
    try:
        with sqlite3.connect("estado.db") as conn:
            miCursor = conn.cursor()
            miCursor.execute("SELECT claveCliente, nombre, apellidos FROM clientes ORDER BY claveCliente ASC;")
            clientesMostrados = miCursor.fetchall()

            if clientesMostrados:
                print("\nClientes registrados:")
                print("*"*30)
                print("Clave\tNombre completo")
                for clave, nombre, apellidos in clientesMostrados:
                    print(f"{clave}\t{nombre} {apellidos}")
                print("*"*30)
            else:
                print("ⓘ No hay clientes registrados.")
    except Error as e:
        print(e)
        return
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        return
    if not clientesMostrados:
        return False
    else:
        return True

def reservarSala():
    if not mostrarClientes():
        return
    while True:
        try:
            claveCliente = int(input("Ingresa tu clave de cliente: "))
        except ValueError:
            print("⚠︎ Clave inválida.")
            continue

        try:
            with sqlite3.connect("estado.db") as conn:
                miCursor = conn.cursor()
                miCursor.execute("SELECT * FROM clientes WHERE claveCliente = ?;", (claveCliente,))
                cliente = miCursor.fetchone()
        except Error as e:
            print(e)
            return
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

        if cliente is None:
            print("⚠︎ La clave de cliente no existe.")
            opcionCancelar = input("¿Cancelar operación? (S/N) ").upper()
            if opcionCancelar == "S":
                return
            elif opcionCancelar == "N":
                continue
            else:
                print("⚠︎ Opción no reconocida.")
                continue
        else:
            clienteAgendado = claveCliente
            break
    while True:
        print("Escribe '0' para cancelar la operacion.")
        fecha_str = input("Ingresa la fecha a agendar (mm-dd-aaaa): ")
        try:
            fechaAgendada = dt.datetime.strptime(fecha_str, "%m-%d-%Y")
        except ValueError:
            print("⚠︎ Fecha inválida.")
            opcionCancelar = input("¿Cancelar operación? (S/N) ").upper()
            if opcionCancelar == "S":
                return
            elif opcionCancelar == "N":
                continue
            else:
                print("⚠︎ Opción no reconocida.")
                continue
        if fechaAgendada.weekday() == 6:
            diaSiguiente = fechaAgendada + dt.timedelta(days=1)
            print(f"ⓘ La reservacion no puede ser hecha en domingo. Deseas agendar el lunes {diaSiguiente.strftime("%d %b %Y")}? (S/N)")
            opcionReagendar = input().upper()
            if opcionReagendar == "S":
                fechaAgendada = diaSiguiente
            elif opcionReagendar == "N":
                continue
            else:
                print("⚠︎ Opción no reconocida.")
                continue
        if fechaAgendada >= (hoy + timedelta(days=2)):
            break
        else:
            print("ⓘ La reservación tiene que ser hecha con 2 (dos) días de anticipación como mínimo.")
            continue

    try:
        with sqlite3.connect("estado.db") as conn:
            miCursor = conn.cursor()
            miCursor.execute("SELECT claveSala, nombre, cupo FROM salas;")
            salas = miCursor.fetchall()

            if salas:
                print(f"\nSALAS Y TURNOS DISPONIBLES EL {fechaAgendada.strftime('%d %b %Y')}:")
                print("*" * 65)
                print(f"{'Clave':<6}\t{'Nombre':<15}\t{'Cupo':<5}\t{'Turnos disponibles':<30}")

                todosTurnos = {"M", "V", "N"}
                disponibilidad = {}

                for claveSala, nombre, cupo in salas:
                    miCursor.execute("SELECT turno FROM reservaciones WHERE sala = ? AND fecha = ?;", (claveSala, fechaAgendada.strftime("%Y-%m-%d")))
                    turnosOcupados = {fila[0] for fila in miCursor.fetchall()}
                    turnosLibres = sorted(todosTurnos - turnosOcupados)
                    disponibilidad[claveSala] = turnosLibres

                    print(f"{claveSala:<6}\t{nombre:<15}\t{cupo:<5}\t{', '.join(turnosLibres) or 'Sin disponibilidad':<30}")

                print("*" * 65)
            else:
                print("ⓘ No hay salas registradas.")
                return
    except Error as e:
        print(f"Error al consultar salas: {e}")
        return
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        return

    while True:
        try:
            salaAgendada = int(input("Ingresa la clave de la sala a agendar: "))
        except ValueError:
            print("⚠︎ Clave inválida.")
            continue

        if salaAgendada == 0:
            return

        if salaAgendada not in disponibilidad.keys():
            print("ⓘ La sala no existe.")
            continue

        if not disponibilidad[salaAgendada]:
            print("ⓘ Esta sala no tiene turnos disponibles en la fecha seleccionada.")
            continue
        break
    while True:
        turno = input("Elige el turno (M - matutino, V - vespertino, N - nocturno): ").upper()
        if turno not in disponibilidad[salaAgendada]:
            print("⚠︎ Turno no disponible.")
            continue
        if turno == "0":
            return
        turnoAgendado = turno
        break
    while True:
        print("Escribe '0' para cancelar la operación.")
        nombreEvento = input("Ingresa el nombre del evento: ")
        if nombreEvento == "" or nombreEvento.strip() == "":
            print("ⓘ El nombre del evento no puede estar vacío.")
            continue
        break

    if nombreEvento == "0":
        return
    else:
        try:
            with sqlite3.connect("estado.db") as conn:
                miCursor = conn.cursor()
                datosReservacion = (fechaAgendada.strftime("%Y-%m-%d"), turnoAgendado, salaAgendada, clienteAgendado, nombreEvento)
                miCursor.execute("INSERT INTO reservaciones (fecha, turno, sala, cliente, nombreEvento) VALUES (?, ?, ?, ?, ?);", datosReservacion)
                print("✓ La reservación fue registrada con éxito.\n")
        except Error as e:
            print(f"Error al registrar la reservación: {e}")
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            return

def exportarXLSX(fecha:dt.datetime):
    fecha_str = dt.datetime.strftime(fecha, '%d %b %Y')
    fecha_str_nums = dt.datetime.strftime(fecha, '%m-%d-%Y')
    wbReporte = pyxl.Workbook()
    hoja = wbReporte["Sheet"]
    hoja.title = "Reservaciones"
    hoja.merge_cells("A1:D1")
    hoja["A1"] = f"RESERVACIONES DEL DIA {fecha_str}"
    encabezados = ["Sala", "Cliente", "Evento", "Turno"]
    hoja.append(encabezados)
    hoja.column_dimensions["A"].width = 10
    hoja.column_dimensions["B"].width = 20
    hoja.column_dimensions["C"].width = 25
    hoja.column_dimensions["D"].width = 10

    try:
        with sqlite3.connect("estado.db") as conn:
            miCursor = conn.cursor()
            miCursor.execute("""
                SELECT s.nombre AS sala,
                       c.nombre || ' ' || c.apellidos AS cliente,
                       r.nombreEvento,
                       r.turno
                FROM reservaciones r
                INNER JOIN salas s ON r.sala = s.claveSala
                INNER JOIN clientes c ON r.cliente = c.claveCliente
                WHERE r.fecha = ?;
            """, (fecha.strftime("%Y-%m-%d"),))

            resultados = miCursor.fetchall()

            for sala, cliente, evento, turno in resultados:
                fila = [sala, cliente, evento, nombresTurnos.get(turno, turno)]
                hoja.append(fila)

    except Error as e:
        print(f"Error al consultar la base de datos: {e}")
        return
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        return

    negritas = Font(bold = True)
    centrado = Alignment(horizontal="center", vertical="center")
    bordeInferior = Border(bottom=Side(border_style="thick", color="000000"))
    hoja["A1"].font = negritas
    for fila in hoja["A2:D2"]:
        for celda in fila:
            celda.font = negritas
            celda.alignment = centrado
            celda.border = bordeInferior
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
        for celda in fila:
            celda.alignment = centrado

    nombreArchivoExportar = f"reporte_{fecha_str_nums}.xlsx"
    wbReporte.save(nombreArchivoExportar)

def consultarReservaciones():
    print("Para consultar una reservación, ingresa la fecha (mm-dd-aaaa) bajo la que fue agendada.")
    print("Escribe '0' en cualquier campo para cancelar la operación.")

    while True:
        fechaConsultada_str = input("Fecha a consultar: ")
        if fechaConsultada_str == "0":
            return
        try:
            fechaConsultada = dt.datetime.strptime(fechaConsultada_str, "%m-%d-%Y")
            fecha_sql = fechaConsultada.strftime("%Y-%m-%d")
        except:
            print("⚠︎ Fecha inválida.")
            continue
        try:
            with sqlite3.connect("estado.db") as conn:
                miCursor = conn.cursor()
                miCursor.execute("""
                    SELECT
                        r.fecha,
                        r.turno,
                        s.nombre AS sala,
                        c.nombre || ' ' || c.apellidos AS cliente,
                        r.nombreEvento
                    FROM reservaciones AS r
                    JOIN salas AS s ON r.sala = s.claveSala
                    JOIN clientes AS c ON r.cliente = c.claveCliente
                    WHERE r.fecha = ?;
                """, (fecha_sql,))

                registros = miCursor.fetchall()

                if not registros:
                    print("ⓘ No existen reservaciones agendadas bajo esta fecha.")
                    opcionCancelar = input("¿Cancelar operación? (S/N) ").upper()
                    if opcionCancelar == "S":
                        return
                    elif opcionCancelar == "N":
                        continue
                    else:
                        print("⚠︎ Opción no reconocida.")
                        continue

                print("\n")
                print("*"*70)
                print(f"REPORTE DE RESERVACIONES PARA EL DÍA {fechaConsultada.strftime('%d %b %Y')}")
                print("*"*70)
                print(f"{'Sala':<10}{'Cliente':<20}{'Evento':<30}{'Turno':<10}")
                print("-"*70)
                for fila in registros:
                    _, turno, sala, cliente, evento = fila
                    print(f"{sala:<10}{cliente:<20}{evento:<30}{turno:<10}")
                print("*"*70)

                while True:
                    opcionExportar = input(f"¿Deseas exportar el reporte del {fechaConsultada.strftime('%d %b %Y')} a Excel? (S/N) ").upper()
                    if opcionExportar == "S":
                        exportarXLSX(fechaConsultada)
                        print("✓ El reporte fue exportado exitosamente.")
                        break
                    elif opcionExportar == "N":
                        break
                    else:
                        print("⚠︎ Opción no reconocida.")
                        continue
                break
        except Error as e:
            print(f"⚠︎ Error al consultar la base de datos: {e}")
            return
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            return

def registrarSala():
    print("Escribe '0' en cualquier campo para cancelar la operación.")
    while True:
        nombreSala = input("Ingresa el nombre de la sala: ")
        if nombreSala == "0":
            return
        if nombreSala == "" or nombreSala.strip() == "":
            print("ⓘ El nombre no puede estar vacío.")
            continue
        break
    while True:
        try:
            cupoSala = int(input("Ingresa el cupo de la sala: "))
        except ValueError:
            print("ⓘ El valor ingresado no es válido. Debe ser un entero.")
            continue
        if cupoSala == 0:
            return
        break
    try:
        with sqlite3.connect("estado.db") as conn:
            miCursor = conn.cursor()
            nuevaSala = (nombreSala, cupoSala)
            miCursor.execute("INSERT INTO salas (nombre, cupo) VALUES (?, ?);", nuevaSala)
            print("✓ La sala fue registrada con éxito.\n")
    except Error as e:
        print(e)
        return
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        return

def editarEvento():
    print("Para editar el nombre de un evento existente, ingresa el rango de fechas (mm-dd-aaaa) en el que se encuentra agendado el evento que quieres editar.")
    print("Escribe '0' en cualquier campo para cancelar la operación.")

    while True:
        inicioRango_str = input("Del: ")
        if inicioRango_str == "0":
            return
        finRango_str = input("Al: ")
        if finRango_str == "0":
            return

        try:
            inicioRango = dt.datetime.strptime(inicioRango_str, "%m-%d-%Y")
            finRango = dt.datetime.strptime(finRango_str, "%m-%d-%Y")
            inicioRango_sql = inicioRango.strftime("%Y-%m-%d")
            finRango_sql = finRango.strftime("%Y-%m-%d")
        except:
            print("⚠︎ Fecha inválida.")
            opcionCancelar = input("¿Cancelar operación? (S/N) ").upper()
            if opcionCancelar == "S":
                return
            elif opcionCancelar == "N":
                continue
            else:
                print("⚠︎ Opción no reconocida.")
                continue

        try:
            with sqlite3.connect("estado.db") as conn:
                miCursor = conn.cursor()
                miCursor.execute("""
                    SELECT
                        r.rowid AS claveEvento,
                        r.fecha,
                        r.nombreEvento,
                        r.turno,
                        s.nombre AS sala
                    FROM reservaciones AS r
                    JOIN salas AS s ON r.sala = s.claveSala
                    WHERE r.fecha BETWEEN ? AND ?
                    ORDER BY r.fecha;
                """, (inicioRango_sql, finRango_sql))

                eventos = miCursor.fetchall()

                if not eventos:
                    print("ⓘ No existen reservaciones en el rango seleccionado.")
                    opcionCancelar = input("¿Cancelar operación? (S/N) ").upper()
                    if opcionCancelar == "S":
                        return
                    elif opcionCancelar == "N":
                        continue
                    else:
                        print("⚠︎ Opción no reconocida.")
                        continue

                print(f"\nEVENTOS REGISTRADOS ENTRE EL {inicioRango.strftime('%d %b %Y')} Y EL {finRango.strftime('%d %b %Y')}:")
                print("*" * 100)
                print(f"{'Clave':<10}{'Fecha':<20}{'Nombre':<30}{'Turno':<20}Sala")
                for evento in eventos:
                    claveEvento, fecha, nombreEvento, turno, sala = evento
                    print(f"{claveEvento:<10}{fecha.strftime("%m-%d-%Y"):<20}{nombreEvento:<30}{turno:<20}{sala}")
                print("*" * 100)

                while True:
                    try:
                        eventoEditando = int(input("\nIngresa la clave del evento que deseas renombrar: "))
                    except ValueError:
                        print("⚠︎ Clave inválida.")
                        continue
                    if eventoEditando == 0:
                        return

                    miCursor.execute("SELECT nombreEvento FROM reservaciones WHERE rowid = ?;", (eventoEditando,))
                    registro = miCursor.fetchone()
                    if not registro:
                        print("ⓘ Este evento no existe en el rango de fechas.")
                        continue
                    break

                while True:
                    nuevoNombre = input("Ingresa el nuevo nombre para el evento: ")
                    if nuevoNombre == "0":
                        return
                    if not nuevoNombre.strip():
                        print("ⓘ El nombre no puede estar vacío.")
                        continue
                    break

                miCursor.execute("UPDATE reservaciones SET nombreEvento = ? WHERE rowid = ?;", (nuevoNombre, eventoEditando))
                print(f"✓ El nombre del evento con clave {eventoEditando} fue editado a '{nuevoNombre}' exitosamente.\n")
                break

        except Error as e:
            print(f"⚠︎ Error al acceder a la base de datos: {e}")
            return
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            return

def crearBD():
    if not os.path.exists("estado.db"):
        try:
            with sqlite3.connect("estado.db") as conn:
                conn.execute("PRAGMA foreign_keys = 1;")
                miCursor = conn.cursor()
                miCursor.execute("CREATE TABLE IF NOT EXISTS salas (claveSala INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL);")
                miCursor.execute("CREATE TABLE IF NOT EXISTS clientes (claveCliente INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellidos TEXT NOT NULL);")
                miCursor.execute("""
                    CREATE TABLE IF NOT EXISTS reservaciones (
                        claveReservacion INTEGER PRIMARY KEY,
                        fecha TEXT NOT NULL,
                        turno TEXT NOT NULL,
                        sala INTEGER NOT NULL,
                        cliente INTEGER NOT NULL,
                        nombreEvento TEXT NOT NULL,
                        FOREIGN KEY (sala) REFERENCES salas(claveSala)
                            ON UPDATE CASCADE
                            ON DELETE RESTRICT,
                        FOREIGN KEY (cliente) REFERENCES clientes(claveCliente)
                            ON UPDATE CASCADE
                            ON DELETE RESTRICT
                    );
                """)
        except Error as e:
            print(e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    else:
        print("ⓘ Los datos de 'estado.db' fueron cargados exitosamente.")

def menu():
    while True:
        print("\n")
        print("*"*50)
        print("SISTEMA DE RESERVA DE SALAS PARA COWORKING")
        print("\nSelecciona una opción para continuar:")
        print("(a) Reservar una sala")
        print("(b) Editar el nombre de una reservación")
        print("(c) Consultar reservaciones")
        print("(d) Registrar nuevo cliente")
        print("(e) Registrar nueva sala")
        print("(f) Salir\n")
        print("*"*50)
        opcion = input()
        if opcion.lower() == "a":
            reservarSala()
            continue
        elif opcion.lower() == "b":
            editarEvento()
            continue
        elif opcion.lower() == "c":
            consultarReservaciones()
            continue
        elif opcion.lower() == "d":
            agregarCliente()
            continue
        elif opcion.lower() == "e":
            registrarSala()
            continue
        elif opcion.lower() == "f":
            opcionSalir = input("¿Guardar y salir? (S/N) ").upper()
            if opcionSalir == "S":
                print("Saliendo...")
                break
            elif opcionSalir == "N":
                continue
            else:
                print("⚠︎ Opción no reconocida.")
                continue

crearBD()
menu()
